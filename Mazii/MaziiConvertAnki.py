from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm

import jaconv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sudachipy import Dictionary, SplitMode
from pykakasi import kakasi
import requests


SOURCE_HEADERS = ["word", "phonetic", "mean", "comment"]
OUTPUT_HEADERS = [
    "Word",
    "Word Reading",
    "Word Meaning",
    "Word Furigana",
    "Vietnamese Sound",
]

DEFAULT_INPUT = Path(__file__).with_name("混合-Mazii.xlsx")
DEFAULT_OUTPUT = Path(__file__).with_name("MaziiConvert_output.xlsx")

MYMEMORY_URL = "https://api.mymemory.translated.net/get"

REQUEST_TIMEOUT = 20

_sudachi = Dictionary().create()
_kakasi = kakasi()

ANSI_GREEN = "\033[92m"
ANSI_RESET = "\033[0m"


class CliTerminated(Exception):
    pass



@dataclass
class ConvertStats:
    rows_seen: int = 0
    rows_written: int = 0
    translation_fallbacks: int = 0
    furigana_fallbacks: int = 0


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_input(prompt: str) -> str:
    try:
        return input(prompt)
    except (KeyboardInterrupt, EOFError):
        raise CliTerminated()


def normalize_reading(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return jaconv.kata2hira(text).lower()


def contains_kanji(text: str) -> bool:
    return bool(re.search(r"[\u4E00-\u9FFF々〆ヶ]", text))


def contains_japanese(text: str) -> bool:
    return bool(re.search(r"[\u4E00-\u9FFF々〆ヶ\u3041-\u3096\u30A1-\u30FA]", text))


def should_fill_vietnamese_sound(word: str, candidate: str) -> bool:
    word_text = normalize_text(word)
    candidate_text = normalize_text(candidate)
    if not candidate_text:
        return False
    if not contains_kanji(word_text) and contains_japanese(word_text):
        return False
    if contains_japanese(candidate_text):
        return False
    return True


@lru_cache(maxsize=4096)
def reading_of(text: str) -> str:
    converted = _kakasi.convert(text)
    if not converted:
        return normalize_reading(text)
    return "".join(normalize_reading(part["hira"]) for part in converted)


def surface_char_kind(char: str) -> str:
    if contains_kanji(char):
        return "kanji"
    if re.fullmatch(r"[\u3041-\u3096\u30A1-\u30FAー]", char):
        return "kana"
    return "other"


def split_surface_runs(text: str) -> list[tuple[str, str]]:
    if not text:
        return []

    runs = []
    start = 0
    current_kind = surface_char_kind(text[0])

    for index, char in enumerate(text[1:], start=1):
        kind = surface_char_kind(char)
        if kind != current_kind:
            runs.append((current_kind, text[start:index]))
            start = index
            current_kind = kind

    runs.append((current_kind, text[start:]))
    return runs


def format_surface_with_reading(surface: str, reading: str) -> str | None:
    runs = split_surface_runs(surface)
    if not runs:
        return None

    @lru_cache(maxsize=None)
    def build(run_index: int, reading_index: int):
        if run_index == len(runs):
            return "" if reading_index == len(reading) else None

        kind, text = runs[run_index]

        if kind == "kana":
            expected = normalize_reading(text)
            if reading.startswith(expected, reading_index):
                tail = build(run_index + 1, reading_index + len(expected))
                if tail is not None:
                    return text + tail
            return None

        if kind == "other":
            tail = build(run_index + 1, reading_index)
            if tail is not None:
                return text + tail
            return None

        for end in range(reading_index + 1, len(reading) + 1):
            tail = build(run_index + 1, end)
            if tail is not None:
                return f"{text}[{reading[reading_index:end]}]" + tail

        return None

    return build(0, 0)


def tokenize_surface(text: str):
    try:
        return list(_sudachi.tokenize(text, SplitMode.A))
    except Exception:
        return []


def format_furigana(text: str) -> str:
    surface = normalize_text(text)
    if not surface:
        return ""
    if not contains_japanese(surface) or not contains_kanji(surface):
        return surface

    tokens = tokenize_surface(surface)
    if len(tokens) > 1:
        pieces = [format_chunk(token.surface()) for token in tokens if token.surface()]
        return "".join(piece for piece in pieces if piece)

    return format_chunk(surface)


def format_chunk(surface: str) -> str:
    surface = normalize_text(surface)
    if not surface:
        return ""

    reading = reading_of(surface)
    if not reading or reading == normalize_reading(surface):
        return surface

    if not contains_kanji(surface):
        return surface

    aligned = format_surface_with_reading(surface, reading)
    if aligned:
        return aligned

    exact_segments = find_exact_segments(surface, reading)
    if exact_segments and len(exact_segments) > 1:
        return " ".join(format_chunk(segment) for segment in exact_segments)

    return f"{surface}[{reading}]"


def find_exact_segments(surface: str, reading: str):
    @lru_cache(maxsize=None)
    def best_from(index: int, reading_index: int):
        if index == len(surface) and reading_index == len(reading):
            return ()
        if index >= len(surface) or reading_index >= len(reading):
            return None

        best = None
        for end in range(index + 1, len(surface) + 1):
            segment = surface[index:end]
            segment_reading = reading_of(segment)
            if not segment_reading:
                continue
            if not reading.startswith(segment_reading, reading_index):
                continue

            remainder = best_from(end, reading_index + len(segment_reading))
            if remainder is None:
                continue

            candidate = ((segment, segment_reading),) + remainder
            if best is None or segment_score(candidate) > segment_score(best):
                best = candidate

        return best

    result = best_from(0, 0)
    if not result:
        return None
    return [segment for segment, _ in result]


def segment_score(segments) -> tuple:
    return (len(segments),) + tuple(len(surface) for surface, _ in segments)


@lru_cache(maxsize=8192)
def translate_vietnamese_to_english(text: str) -> tuple[str, bool]:
    source = normalize_text(text)
    if not source:
        return "", False

    try:
        response = requests.get(
            MYMEMORY_URL,
            params={"q": source, "langpair": "vi|en"},
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        payload = response.json()
        translated = normalize_text(payload.get("responseData", {}).get("translatedText"))
        if translated:
            return translated, translated != source
    except Exception:
        pass

    return source, False


def read_source_rows(source_path: Path):
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [normalize_text(value).lower() for value in rows[0]]
    index_map = {name: position for position, name in enumerate(header) if name}
    has_known_header = all(name in index_map for name in SOURCE_HEADERS)

    if has_known_header:
        data_rows = rows[1:]
    else:
        data_rows = rows
        index_map = {name: position for position, name in enumerate(SOURCE_HEADERS)}

    extracted = []
    for row in data_rows:
        extracted.append({name: normalize_text(row[index_map.get(name, -1)]) if index_map.get(name, -1) >= 0 and index_map.get(name, -1) < len(row) else "" for name in SOURCE_HEADERS})

    return extracted


def convert_rows(rows: Iterable[dict]) -> tuple[list[dict], ConvertStats]:
    row_list = list(rows)
    if not row_list:
        return [], ConvertStats()

    worker_count = min(8, max(1, len(row_list)))
    converted_rows = []
    stats = ConvertStats(rows_seen=len(row_list), rows_written=len(row_list))
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        for converted_row, row_stats in tqdm(
            executor.map(convert_single_row, row_list),
            total=len(row_list),
            desc="Converting rows",
            unit="row",
            dynamic_ncols=True,
        ):
            converted_rows.append(converted_row)
            stats.translation_fallbacks += row_stats.translation_fallbacks
            stats.furigana_fallbacks += row_stats.furigana_fallbacks

    tqdm.write(f"{ANSI_GREEN}Completed converting {len(row_list)} rows.{ANSI_RESET}")
    return converted_rows, stats


def convert_single_row(row: dict) -> tuple[dict, ConvertStats]:
    row_stats = ConvertStats(rows_seen=1, rows_written=1)
    word = normalize_text(row.get("word"))
    phonetic = normalize_reading(row.get("phonetic"))
    mean = normalize_text(row.get("mean"))
    comment = normalize_text(row.get("comment"))

    translated_mean, translated = translate_vietnamese_to_english(mean)
    if not translated and mean:
        row_stats.translation_fallbacks += 1

    furigana = format_furigana(word)
    if furigana and "[" not in furigana and contains_kanji(word):
        row_stats.furigana_fallbacks += 1

    vietnamese_sound = comment.upper() if should_fill_vietnamese_sound(word, comment) else ""

    return (
        {
            "Word": word,
            "Word Reading": phonetic,
            "Word Meaning": translated_mean,
            "Word Furigana": furigana,
            "Vietnamese Sound": vietnamese_sound,
        },
        row_stats,
    )


def write_output_workbook(rows: Iterable[dict], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MaziiConvert"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for column, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        1: 20,
        2: 20,
        3: 42,
        4: 36,
        5: 28,
    }

    for column_index, width in widths.items():
        sheet.column_dimensions[chr(64 + column_index)].width = width

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(OUTPUT_HEADERS, start=1):
            sheet.cell(row_index, column_index, row.get(header, ""))

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def write_source_workbook(rows: Iterable[dict], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MaziiSplit"

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    for column, header in enumerate(SOURCE_HEADERS, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        1: 24,
        2: 24,
        3: 42,
        4: 28,
    }

    for column_index, width in widths.items():
        sheet.column_dimensions[chr(64 + column_index)].width = width

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(SOURCE_HEADERS, start=1):
            sheet.cell(row_index, column_index, row.get(header, ""))

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def run_full_conversion(input_path: Path, output_path: Path) -> ConvertStats:
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    source_rows = read_source_rows(input_path)
    converted_rows, stats = convert_rows(source_rows)
    write_output_workbook(converted_rows, output_path)

    print(f"Converted {stats.rows_written} rows -> {output_path}")
    print(f"Translation fallbacks: {stats.translation_fallbacks}")
    print(f"Furigana fallbacks: {stats.furigana_fallbacks}")
    return stats


def preview_source(input_path: Path, count: int = 5) -> None:
    rows = read_source_rows(input_path)
    for index, row in enumerate(rows[:count], start=1):
        print(
            f"{index}. word={row.get('word', '')} | phonetic={row.get('phonetic', '')} | mean={row.get('mean', '')} | comment={row.get('comment', '')}"
        )


def should_skip_input_file(path: Path) -> bool:
    name_lower = path.name.lower()
    stem_lower = path.stem.lower()
    if name_lower == DEFAULT_OUTPUT.name.lower():
        return True
    if stem_lower.endswith("_output"):
        return True
    if re.search(r"_part\d+$", stem_lower):
        return True
    if stem_lower.endswith("_single_kanji"):
        return True
    if stem_lower.endswith("_kanji_only"):
        return True
    if stem_lower.endswith("_katakana_only"):
        return True
    if stem_lower.endswith("_others"):
        return True
    return False


def is_single_kanji(text: str) -> bool:
    value = normalize_text(text)
    return len(value) == 1 and bool(re.fullmatch(r"[\u4E00-\u9FFF々〆ヶ]", value))


def is_kanji_only(text: str) -> bool:
    value = normalize_text(text)
    return bool(value) and bool(re.fullmatch(r"[\u4E00-\u9FFF々〆ヶ]+", value))


def is_katakana_only(text: str) -> bool:
    value = normalize_text(text)
    return bool(value) and bool(re.fullmatch(r"[\u30A0-\u30FFー]+", value))


def classify_word_type(word: str) -> str:
    if is_single_kanji(word):
        return "single_kanji"
    if is_kanji_only(word):
        return "kanji_only"
    if is_katakana_only(word):
        return "katakana_only"
    return "others"


def discover_input_files(workdir: Path) -> list[Path]:
    files = []
    for path in sorted(workdir.glob("*.xlsx")):
        if path.is_file() and not should_skip_input_file(path):
            files.append(path.resolve())
    return files


def output_path_for_input(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_output.xlsx")


def pick_one_file(files: list[Path]) -> Path | None:
    if not files:
        print("No eligible .xlsx files found in current directory.")
        return None

    print("\nSelect one file:")
    for idx, file_path in enumerate(files, start=1):
        print(f"{idx}. {file_path.name}")

    while True:
        selected = safe_input("Enter file number (empty to cancel): ").strip()
        if not selected:
            return None
        if not selected.isdigit():
            print("Invalid number. Try again.")
            continue

        index = int(selected)
        if 1 <= index <= len(files):
            return files[index - 1]
        print("Number out of range. Try again.")


def prompt_batch_size(default_size: int = 500) -> int:
    while True:
        value = safe_input(f"Batch size (empty = {default_size}): ").strip()
        if not value:
            return default_size
        if not value.isdigit():
            print("Batch size must be a positive integer.")
            continue
        size = int(value)
        if size <= 0:
            print("Batch size must be greater than 0.")
            continue
        return size


def menu_loop(input_path: Path, output_path: Path) -> None:
    while True:
        workdir = Path.cwd().resolve()
        files = discover_input_files(workdir)
        print("\nMaziiConvert CLI")
        print(f"Directory: {workdir}")
        print(f"Found .xlsx inputs: {len(files)}")
        print("1. Convert")
        print("2. Split files")
        print("3. Split by word type")
        print("4. Quit")

        choice = safe_input("Select an option: ").strip()

        if choice == "1":
            if not files:
                print("No eligible .xlsx files found in current directory.")
                continue

            print("\nConvert")
            print("1. All files")
            print("2. Select one file")
            convert_choice = safe_input("Select convert mode: ").strip()

            if convert_choice == "1":
                for file_path in files:
                    out_path = output_path_for_input(file_path)
                    print(f"\nConverting {file_path.name} -> {out_path.name}")
                    run_full_conversion(file_path, out_path)
            elif convert_choice == "2":
                selected = pick_one_file(files)
                if selected is None:
                    continue
                out_path = output_path_for_input(selected)
                print(f"\nConverting {selected.name} -> {out_path.name}")
                run_full_conversion(selected, out_path)
            else:
                print("Invalid choice. Try again.")
        elif choice == "2":
            if not files:
                print("No eligible .xlsx files found in current directory.")
                continue

            batch_size = prompt_batch_size(default_size=500)
            for file_path in tqdm(
                files,
                total=len(files),
                desc="Split files",
                unit="file",
                dynamic_ncols=True,
            ):
                out_prefix = file_path.with_suffix("")
                split_source_batches(file_path, out_prefix, batch_size)
            tqdm.write(f"{ANSI_GREEN}Completed splitting {len(files)} file(s).{ANSI_RESET}")
        elif choice == "3":
            if not files:
                print("No eligible .xlsx files found in current directory.")
                continue

            for file_path in tqdm(
                files,
                total=len(files),
                desc="Split by type",
                unit="file",
                dynamic_ncols=True,
            ):
                split_by_word_type(file_path)
            tqdm.write(f"{ANSI_GREEN}Completed split by word type for {len(files)} file(s).{ANSI_RESET}")
        elif choice == "4":
            return
        else:
            print("Invalid choice. Try again.")


def parse_args(argv: list[str]):
    parser = argparse.ArgumentParser(description="MaziiConvert workbook converter")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input workbook path")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output workbook path")
    parser.add_argument("--full-run", action="store_true", help="Run conversion immediately without the menu")
    return parser.parse_args(argv)


def split_source_batches(input_path: Path, output_prefix: Path, batch_size: int = 500) -> None:
    rows = read_source_rows(input_path)
    if not rows:
        print("No rows found to split.")
        return

    total = len(rows)
    parts = (total + batch_size - 1) // batch_size
    print(f"Splitting {total} rows into {parts} batch(es) of up to {batch_size} rows...")

    for i in tqdm(
        range(parts),
        total=parts,
        desc=f"Splitting {input_path.name}",
        unit="batch",
        dynamic_ncols=True,
    ):
        start = i * batch_size
        end = min(start + batch_size, total)
        chunk = rows[start:end]
        out_path = output_prefix.with_name(f"{output_prefix.stem}_part{i+1}{output_prefix.suffix or '.xlsx'}")
        write_source_workbook(chunk, out_path)
        tqdm.write(f"Wrote {out_path} ({len(chunk)} rows)")

    tqdm.write(f"{ANSI_GREEN}Completed {input_path.name}: {parts} batch(es).{ANSI_RESET}")


def split_by_word_type(input_path: Path) -> None:
    rows = read_source_rows(input_path)
    if not rows:
        print(f"No rows found in {input_path.name} to split by word type.")
        return

    buckets: dict[str, list[dict]] = {
        "single_kanji": [],
        "kanji_only": [],
        "katakana_only": [],
        "others": [],
    }

    for row in tqdm(rows, total=len(rows), desc=f"Classify {input_path.name}", unit="row", dynamic_ncols=True):
        word = normalize_text(row.get("word"))
        bucket = classify_word_type(word)
        buckets[bucket].append(row)

    for bucket_name, bucket_rows in buckets.items():
        if not bucket_rows:
            continue
        out_path = input_path.with_name(f"{input_path.stem}_{bucket_name}.xlsx")
        write_source_workbook(bucket_rows, out_path)
        tqdm.write(f"Wrote {out_path} ({len(bucket_rows)} rows)")

    tqdm.write(f"{ANSI_GREEN}Completed type split for {input_path.name}.{ANSI_RESET}")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    input_path = args.input.expanduser().resolve()
    output_path = args.output.expanduser().resolve()

    try:
        if args.full_run:
            run_full_conversion(input_path, output_path)
        else:
            menu_loop(input_path, output_path)
    except CliTerminated:
        print("\nCLI terminated by user.")
        return 0
    except KeyboardInterrupt:
        print("\nCLI terminated by user.")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

